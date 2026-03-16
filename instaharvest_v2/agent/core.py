"""
InstaAgent — Core AI Agent
===========================
Main agent class that connects LLM + InstaHarvest v2 + Executor + Permissions.

Usage:
    from instaharvest_v2 import Instagram
    from instaharvest_v2.agent import InstaAgent, Permission

    ig = Instagram.from_env(".env")

    # With OpenAI
    agent = InstaAgent(ig, provider="openai", api_key="sk-...")

    # With Gemini
    agent = InstaAgent(ig, provider="gemini", api_key="...")

    # Ask a question
    result = agent.ask("How many followers does Cristiano have?")
    print(result)

    # Interactive chat
    agent.chat()

    # Full access (no permission prompts)
    agent = InstaAgent(ig, provider="gemini", api_key="...",
                       permission=Permission.FULL_ACCESS)
"""

import json
import logging
import os
import time
from dataclasses import dataclass, field
from typing import Any, Callable, Dict, List, Optional

from .knowledge import SYSTEM_PROMPT
from .executor import SafeExecutor, ExecutionResult
from .permissions import Permission, PermissionManager, ActionType
from .providers import get_provider, resolve_api_key
from .providers.base import BaseProvider, ProviderResponse, ToolCall
from .tools import TOOL_HANDLERS
from .memory import AgentMemory
from .plugins import PluginManager
from .cost_tracker import CostTracker
from .retry import RetryPolicy
from .streaming import StreamHandler
from .compat import (
    safe_print, emoji, setup_console_encoding,
    get_default_memory_dir, get_default_cost_path,
    get_platform_info,
)

logger = logging.getLogger("instaharvest_v2.agent")


@dataclass
class AgentResult:
    """Result of an agent operation."""
    answer: str = ""
    code_executed: str = ""
    execution_result: Optional[ExecutionResult] = None
    files_created: List[str] = field(default_factory=list)
    steps: int = 0
    tokens_used: int = 0
    duration: float = 0.0
    error: str = ""

    @property
    def success(self) -> bool:
        return not self.error

    def __str__(self) -> str:
        if self.error:
            return f"❌ {self.error}"
        return self.answer or "Completed"


class InstaAgent:
    """
    AI Agent for instaharvest_v2.

    Connects any LLM with InstaHarvest v2 library.
    User speaks naturally, agent writes and executes code.

    Supported modes:
        - Login: Instagram.from_env(".env") — full API access
        - Anonymous: Instagram() or ig=None — only ig.public.*
        - Async: AsyncInstagram — async wrapper

    Supported providers:
        openai, gemini, claude, deepseek, qwen, groq, together,
        mistral, ollama, openrouter, fireworks, perplexity, xai, custom

    Args:
        ig: Instagram instance (None = anonymous mode)
        provider: AI provider name (default: 'gemini')
        api_key: AI provider API key (or auto-detect from env)
        model: Optional model override
        permission: Permission level (ASK_EVERY, ASK_ONCE, FULL_ACCESS)
        permission_callback: Custom permission prompt function
        max_steps: Maximum agent loop iterations (default 15)
        timeout: Code execution timeout in seconds (default 30)
        verbose: Print step-by-step progress
    """

    def __init__(
        self,
        ig=None,
        provider: str = "gemini",
        api_key: Optional[str] = None,
        model: Optional[str] = None,
        permission: Permission = Permission.ASK_EVERY,
        permission_callback: Optional[Callable] = None,
        max_steps: int = 15,
        timeout: int = 30,
        verbose: bool = True,
        memory: bool = False,
        memory_dir: Optional[str] = None,
        cost_tracking: bool = True,
        retry_count: int = 2,
        streaming: bool = False,
    ):
        # Resolve API key
        api_key = api_key or resolve_api_key(provider)
        if not api_key and provider.lower() != "ollama":
            raise ValueError(
                f"API key not found for '{provider}'!\n"
                f"  1. Pass api_key='...' parameter\n"
                f"  2. Add the key to .env file\n"
                f"  Supported: openai, gemini, claude, deepseek, "
                f"qwen, groq, together, mistral, ollama, openrouter, xai"
            )

        self._ig = ig
        self._provider: BaseProvider = get_provider(provider, api_key, model)
        self._model_name = model or self._provider.model_name if hasattr(self._provider, 'model_name') else 'auto'
        self._executor = SafeExecutor(ig_instance=ig, timeout=timeout)
        self._permissions = PermissionManager(
            level=permission,
            prompt_callback=permission_callback,
        )
        self._max_steps = max_steps
        self._verbose = verbose
        self._history: List[Dict[str, Any]] = []
        self._files_created: List[str] = []

        # User data cache — persists across agent calls
        self._user_cache: Dict[str, Any] = {}

        # ── Pro Architecture ──────────────────────────────
        # Dead loop detection: track consecutive identical errors
        self._error_tracker: Dict[str, int] = {}
        self._max_identical_errors = 2

        # Tool usage analytics
        self._tool_stats: Dict[str, int] = {}
        self._total_queries = 0

        # Audit trail
        self._audit_enabled = True
        self._audit_dir = os.path.join(
            os.path.expanduser("~"), ".instaharvest_v2", "agent_logs"
        )

        # Advanced features
        self._memory = AgentMemory(memory_dir=memory_dir) if memory else None
        self._plugins = PluginManager()
        self._cost_tracker = CostTracker() if cost_tracking else None
        self._retry = RetryPolicy(max_retries=retry_count)
        self._stream = StreamHandler(mode="cli") if streaming else None
        self._session_id = AgentMemory.generate_session_id() if memory else ""

        # Detect mode — check if actually logged in
        self._is_logged_in = False
        ig_class = type(ig).__name__ if ig else "None"
        if ig is None:
            self._mode = "anonymous"
        elif "Async" in ig_class:
            self._mode = "async"
            self._is_logged_in = True
        else:
            self._mode = "sync"
            # Check if ig actually has sessions (cookies)
            try:
                if hasattr(ig, '_session_mgr') and ig._session_mgr.session_count > 0:
                    self._is_logged_in = True
                else:
                    self._is_logged_in = False
            except Exception:
                self._is_logged_in = False

        # Inject cache and login status into executor
        self._executor._user_cache = self._user_cache
        self._executor._is_logged_in = self._is_logged_in

        # Build system prompt with detailed mode info
        self._is_local_provider = provider.lower() in ("ollama", "lmstudio", "local", "lm-studio", "lm_studio", "llmstudio")

        if self._is_local_provider:
            # Compact prompt for local models (limited context window)
            system_content = self._build_compact_prompt()
        else:
            mode_info = self._build_mode_info()
            # Dynamic API reference — auto-discover available modules/methods
            api_ref = self._build_api_reference()
            system_content = SYSTEM_PROMPT + mode_info + api_ref

        # Initialize with system prompt
        self._history.append({
            "role": "system",
            "content": system_content,
        })

        mode_emoji = {"sync": "🔑", "anonymous": "👤", "async": "⚡"}
        logger.info(
            f"Agent ready | Provider: {self._provider.provider_name} | "
            f"Mode: {self._mode} {mode_emoji.get(self._mode, '')} | "
            f"Permission: {permission.value}"
        )

        if self._verbose:
            setup_console_encoding()
            mode_text = {"sync": "Login", "anonymous": "Anonymous", "async": "Async"}
            safe_print(
                f"{emoji('🤖', '[BOT]')} Agent: {self._provider.provider_name} | "
                f"Mode: {mode_text.get(self._mode, self._mode)}"
            )

    # ═══════════════════════════════════════════════════════════
    # MODE DETECTION & SMART CONTEXT
    # ═══════════════════════════════════════════════════════════

    def _build_compact_prompt(self) -> str:
        """Build a compact system prompt for local models with limited context."""
        mode = "anonymous" if not self._is_logged_in else "logged_in"
        parts = [
            "You are InstaHarvest v2 Agent — an Instagram analytics assistant.",
            "You write Python code using the `ig` object to interact with Instagram.",
            "",
            "RULES:",
            "1. Write executable Python code in ```python blocks",
            "2. Use try/except for all API calls",
            "3. Use f-strings for formatted output",
            "4. Cache results in `_cache` dict",
            "5. NEVER use nested quotes in f-strings",
            "",
        ]

        if mode == "anonymous":
            parts.extend([
                "MODE: ANONYMOUS — only ig.public.* methods work",
                "Main methods:",
                "- ig.public.get_profile(username) → dict",
                "- ig.public.get_posts(username, max_count=12) → list",
                "- ig.public.get_post_by_url(url) → dict",
                "- ig.public.get_comments(shortcode) → list",
                "- ig.public.get_hashtag_posts(tag) → list",
                "",
                "Profile fields: username, full_name, biography, is_verified, is_private,",
                "  followers (count), following (count), posts_count (count),",
                "  profile_pic_url, profile_pic_url_hd",
                "",
                "IMPORTANT: get_profile() returns a flat dict with these TOP-LEVEL keys:",
                "  followers, following, posts_count — use them DIRECTLY!",
                "  Do NOT use edge_followed_by or follower_count — they don't exist!",
                "",
                "Example:",
                "```python",
                "profile = ig.public.get_profile('username')",
                "if profile:",
                "    print(f'Followers: {profile.get(\"followers\", 0):,}')",
                "    print(f'Following: {profile.get(\"following\", 0):,}')",
                "    print(f'Posts: {profile.get(\"posts_count\", 0):,}')",
                "```",
            ])
        else:
            parts.extend([
                "MODE: LOGGED IN — all API methods available",
                "Main methods:",
                "- ig.users.get_by_username(username) → User model",
                "- ig.public.get_profile(username) → dict (fallback)",
                "- ig.feed.user_feed(user_id) → list",
                "- ig.friendships.follow/unfollow(user_id)",
                "- ig.friendships.get_followers/get_following(user_id)",
                "- ig.direct.send_text(thread_id, text)",
                "- ig.media.get_info(media_pk)",
                "",
                "User model: .username, .followers, .following, .posts_count, .biography",
            ])

        return "\n".join(parts)

    def _build_mode_info(self) -> str:
        """Build detailed mode-specific system prompt addition."""
        parts = ["\n\n# CURRENT SESSION STATUS"]

        if self._mode == "anonymous" or not self._is_logged_in:
            parts.append("""
## MODE: ANONYMOUS (No Login) — `_is_logged_in = False`
- `ig` is available, but no cookie/session
- Only `ig.public.*` methods work
- `ig.users.*`, `ig.feed.*`, `ig.friendships.*` — WILL NOT WORK!

### Rules:
1. ALWAYS use `ig.public.get_profile(username)` — it returns a FLAT dict
2. Use `.get()` with correct keys: `profile.get('followers', 0)` — NOT 'follower_count'!
3. NEVER try `ig.users.get_by_username()` — it WILL FAIL, do NOT waste a step!
4. Follow/unfollow/DM/upload — NOT POSSIBLE, tell the user "login required"
5. For a simple profile query: ONE code execution should be enough!
6. Do NOT write "trying login API first..." — go DIRECTLY to ig.public.*

### Example pattern (ANONYMOUS — ONE STEP!):
```python
try:
    profile = ig.public.get_profile('username')
    if profile:
        print(f"Username: {profile.get('username', 'N/A')}")
        print(f"Full Name: {profile.get('full_name', 'N/A')}")
        print(f"Followers: {profile.get('followers', 0):,}")
        print(f"Following: {profile.get('following', 0):,}")
        print(f"Posts: {profile.get('posts_count', 0):,}")
        print(f"Bio: {profile.get('biography', '')}")
        verified = 'Yes' if profile.get('is_verified') else 'No'
        private = 'Yes' if profile.get('is_private') else 'No'
        print(f"Verified: {verified}")
        print(f"Private: {private}")
        _cache[profile.get('username', '')] = profile
    else:
        print("User not found")
except Exception as e:
    print(f"Error: {e}")
```

### IMPORTANT: get_profile() returns a FLAT dict!
- Use `profile.get('followers', 0)` — NOT `profile.get('follower_count')` or `profile.get('edge_followed_by')`
- Use `profile.get('following', 0)` — NOT `profile.get('edge_follow')`
- Use `profile.get('posts_count', 0)` — NOT `profile.get('edge_owner_to_timeline_media')`
- These are the ONLY correct field names. All other field names are WRONG.

### NOT AVAILABLE IN ANONYMOUS MODE:
- ❌ ig.users.get_by_username() — requires session
- ❌ ig.friendships.follow/unfollow — requires login
- ❌ ig.direct.send_text — requires login
- ❌ ig.media.like/comment — requires login
- ❌ ig.upload.* — requires login
- ❌ ig.account.* — requires login
- ❌ ig.stories.get_tray — requires login

### AVAILABLE IN ANONYMOUS MODE:
- ✅ ig.public.get_profile(username) → dict
- ✅ ig.public.get_posts(username) → list
- ✅ ig.public.get_post_by_url(url) → dict
""")
        elif self._mode == "sync" and self._is_logged_in:
            parts.append("""
## MODE: LOGGED IN (Session available ✅)
- All API methods are available
- ig.users.*, ig.feed.*, ig.friendships.*, ig.direct.* — everything!

### Strategy:
1. First use `ig.users.get_by_username()` — returns User model
2. If unsuccessful, use `ig.public.get_profile()` as fallback
3. Save retrieved data to `_cache` — retrieve from cache on next request

### Example pattern (LOGIN):
```python
try:
    # Check cache first
    if 'username' in _cache:
        user = _cache['username']
        print(f"(From cache) Followers: {user.followers:,}")
    else:
        user = ig.users.get_by_username('username')
    print(f"Username: {user.username}")
    print(f"Followers: {user.followers:,}")
    print(f"Following: {user.following:,}")
    print(f"Posts: {user.posts_count}")
    _cache[user.username] = user  # Save to cache
except Exception as e:
    # Fallback to public API
    try:
        profile = ig.public.get_profile('username')
        print(f"Followers: {profile.get('followers', 0):,}")
        _cache['username'] = profile
    except Exception as e2:
        print(f"Error: {e2}")
```

### ADDITIONAL CAPABILITIES IN LOGIN MODE:
- ✅ ig.friendships.follow/unfollow
- ✅ ig.direct.send_text
- ✅ ig.media.like/comment
- ✅ ig.upload.post_photo/video/reel
- ✅ ig.account.get_current_user
- ✅ ig.stories.get_tray
""")
        elif self._mode == "async":
            parts.append("""
## MODE: ASYNC (AsyncInstagram)
- All methods require `await`
- DO NOT use `import asyncio` and `asyncio.run()`!

### Example:
```python
user = await ig.users.get_by_username('username')
print(f"Followers: {user.followers:,}")
```
""")

        # Add cache info
        parts.append("""
# DATA CACHING
- `_cache` dict — store retrieved data here
- `_is_logged_in` — True/False, check login status
- Always check the cache first
Example: if 'cristiano' in _cache: user = _cache['cristiano']
""")

        return "\n".join(parts)

    def _build_api_reference(self) -> str:
        """Dynamically introspects the ig object to build a live API reference.

        This ensures the agent always knows exactly what modules and methods
        are available at runtime, even if the library is updated.
        """
        if self._ig is None:
            return ""

        import inspect

        parts = ["\n\n# LIVE API REFERENCE (auto-discovered from your ig instance)"]
        parts.append("The following modules and methods are available on the `ig` object right now:\n")

        # Known sub-module attribute names to scan
        _skip = {
            "_session_mgr", "_client", "_anon", "_log",
            "__class__", "__dict__", "__weakref__",
        }

        ig = self._ig
        modules_found = []

        for attr_name in sorted(dir(ig)):
            if attr_name.startswith("_") or attr_name in _skip:
                continue
            try:
                module = getattr(ig, attr_name, None)
            except Exception:
                continue
            if module is None or isinstance(module, (str, int, float, bool, list, dict)):
                continue
            # Must be a sub-module object (has methods)
            if not hasattr(module, "__class__"):
                continue

            methods = []
            for method_name in sorted(dir(module)):
                if method_name.startswith("_"):
                    continue
                try:
                    method = getattr(module, method_name, None)
                except Exception:
                    continue
                if not callable(method):
                    continue

                # Get signature
                try:
                    sig = inspect.signature(method)
                    params = []
                    for pname, param in sig.parameters.items():
                        if pname == "self":
                            continue
                        if param.default is inspect.Parameter.empty:
                            params.append(pname)
                        else:
                            params.append(f"{pname}={param.default!r}")
                    sig_str = f"({', '.join(params)})"
                except (ValueError, TypeError):
                    sig_str = "(...)"

                methods.append(f"  - ig.{attr_name}.{method_name}{sig_str}")

            if methods:
                modules_found.append(f"\n## ig.{attr_name}")
                modules_found.extend(methods)

        if modules_found:
            parts.extend(modules_found)
        else:
            parts.append("(No modules discovered — ig may not be fully initialized)")

        return "\n".join(parts)

    @property
    def is_logged_in(self) -> bool:
        """Check if agent has active login session."""
        return self._is_logged_in

    @property
    def user_cache(self) -> Dict:
        """Access agent's user data cache."""
        return self._user_cache

    # ═══════════════════════════════════════════════════════════
    # PUBLIC API
    # ═══════════════════════════════════════════════════════════

    def register_tool(
        self,
        name: str,
        handler: Callable,
        description: str = "",
        schema: Optional[Dict] = None,
    ):
        """Register a custom plugin tool."""
        self._plugins.register(name, handler, description, schema)

    @property
    def memory(self) -> Optional[AgentMemory]:
        """Access agent memory."""
        return self._memory

    @property
    def cost(self) -> Optional[CostTracker]:
        """Access cost tracker."""
        return self._cost_tracker

    @property
    def plugins(self) -> PluginManager:
        """Access plugin manager."""
        return self._plugins

    @property
    def provider_name(self) -> str:
        """Current provider name."""
        return self._provider.provider_name

    @property
    def mode(self) -> str:
        """Current mode (sync/anonymous/async)."""
        return self._mode

    # ═══════════════════════════════════════════════════════════
    # AUDIT TRAIL & ANALYTICS
    # ═══════════════════════════════════════════════════════════

    def _save_audit_log(self, query: str, result) -> None:
        """
        Save audit log for every agent interaction.

        Logs are saved to ~/.instaharvest_v2/agent_logs/YYYY-MM-DD_HHMMSS_NNN.json
        Contains: query, answer, tools called, duration, steps, errors.
        """
        if not self._audit_enabled:
            return

        try:
            os.makedirs(self._audit_dir, exist_ok=True)

            from datetime import datetime
            timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
            log_file = os.path.join(
                self._audit_dir,
                f"{timestamp}_{self._total_queries:03d}.json",
            )

            # Extract tool calls from history
            tools_called = []
            for msg in self._history:
                if msg.get("role") == "tool":
                    tools_called.append({
                        "name": msg.get("name", "?"),
                        "output_preview": str(msg.get("content", ""))[:200],
                    })

            import json
            log_data = {
                "timestamp": timestamp,
                "query": query[:500],
                "answer": str(getattr(result, "answer", ""))[:1000],
                "success": getattr(result, "success", False),
                "steps": getattr(result, "steps", 0),
                "duration": round(getattr(result, "duration", 0), 2),
                "tools_called": tools_called,
                "tool_stats": dict(self._tool_stats),
                "provider": self._provider.provider_name,
                "mode": self._mode,
                "is_logged_in": self._is_logged_in,
                "errors": [
                    k for k, v in self._error_tracker.items() if v > 0
                ],
                "files_created": getattr(result, "files_created", []),
            }

            with open(log_file, "w", encoding="utf-8") as f:
                json.dump(log_data, f, indent=2, ensure_ascii=False)

            logger.debug(f"Audit log saved: {log_file}")

        except Exception as e:
            logger.warning(f"Audit log save failed: {e}")

    @property
    def stats(self) -> Dict[str, Any]:
        """
        Agent analytics: tool usage, query count, top tools.

        Usage:
            agent.stats  →  {
                'total_queries': 5,
                'total_tool_calls': 23,
                'top_tools': [('get_profile', 8), ('write_file', 3)],
                'unique_tools_used': 7,
                'provider': 'gemini',
                'mode': 'sync'
            }
        """
        sorted_tools = sorted(
            self._tool_stats.items(),
            key=lambda x: x[1],
            reverse=True,
        )
        return {
            "total_queries": self._total_queries,
            "total_tool_calls": sum(self._tool_stats.values()),
            "top_tools": sorted_tools[:10],
            "unique_tools_used": len(self._tool_stats),
            "provider": self._provider.provider_name,
            "mode": self._mode,
            "is_logged_in": self._is_logged_in,
        }

    def ask(
        self,
        message: str,
        step_callback: Optional[Callable[[Dict[str, Any]], None]] = None,
    ) -> AgentResult:
        """
        Ask the agent a question or give a command.

        Args:
            message: Natural language message
            step_callback: Optional callback for real-time step events.
                Receives dicts like {"type": "status", "message": "..."}
                Event types: status, thinking, code, tool_call, tool_result, error

        Returns:
            AgentResult with answer and execution details
        """
        start = time.time()

        if self._verbose:
            _e = emoji('\U0001f916', '[BOT]')
            safe_print(f"\n{_e} Agent working...")

        # Notify: starting
        if step_callback:
            step_callback({"type": "status", "message": "Processing your request..."})

        # Compress history if too long (token optimization)
        self._compress_history()

        # Clear error tracker for new query
        self._error_tracker.clear()
        self._total_queries += 1

        # Add user message
        self._history.append({"role": "user", "content": message})

        # Run agent loop
        result = self._agent_loop(step_callback=step_callback)
        result.duration = time.time() - start
        result.tokens_used = self._provider.total_tokens

        if self._verbose and result.success:
            _e = emoji('\u2705', '[OK]')
            safe_print(f"\n{_e} Done ({result.duration:.1f}s, {result.steps} steps)")

        # Save audit log
        self._save_audit_log(message, result)

        return result

    def chat(self) -> None:
        """
        Interactive chat mode.
        Type 'exit' or 'quit' to exit.
        """
        print("\n" + "\u2550" * 50)
        _e = emoji('\U0001f916', '[BOT]')
        safe_print(f"{_e} InstaHarvest v2 Agent \u2014 Interactive mode")
        print(f"   Provider: {self._provider.provider_name}")
        print(f"   Permission: {self._permissions.level.value}")
        print("   'exit' \u2014 quit | 'reset' \u2014 new conversation")
        print("\u2550" * 50)

        while True:
            try:
                user_input = input("\n👤 You: ").strip()
            except (KeyboardInterrupt, EOFError):
                print("\n👋 Goodbye!")
                break

            if not user_input:
                continue
            if user_input.lower() in ("exit", "quit", "q"):
                print("👋 Goodbye!")
                break
            if user_input.lower() in ("reset", "new", "clear"):
                self.reset()
                safe_print(f"{emoji('🔄', '[*]')} Conversation reset")
                continue
            if user_input.lower() in ("history",):
                self._print_history()
                continue

            result = self.ask(user_input)

            if result.success:
                safe_print(f"\n{emoji('🤖', '[BOT]')} Agent: {result.answer}")
            else:
                safe_print(f"\n{emoji('❌', '[ERR]')} Error: {result.error}")

    def reset(self) -> None:
        """Reset conversation history."""
        if self._is_local_provider:
            system_content = self._build_compact_prompt()
        else:
            mode_info = self._build_mode_info()
            api_ref = self._build_api_reference()
            system_content = SYSTEM_PROMPT + mode_info + api_ref
        self._history = [{"role": "system", "content": system_content}]
        self._files_created = []
        self._permissions.reset()
        self._error_tracker.clear()
        logger.info("🔄 Agent reset")

    @property
    def history(self) -> List[Dict]:
        """Conversation history."""
        return [m for m in self._history if m["role"] != "system"]

    @property
    def provider_name(self) -> str:
        return self._provider.provider_name

    # ═══════════════════════════════════════════════════════════
    # DYNAMIC TOOL FILTERING — TOKEN OPTIMIZATION
    # ═══════════════════════════════════════════════════════════

    # Tools that require login — filtered out in anonymous mode
    _LOGIN_REQUIRED_TOOLS = {
        # Upload
        "upload_photo", "upload_video", "upload_reel",
        "upload_story_photo", "upload_story_video", "upload_carousel",
        "delete_media",
        # Media interaction
        "like_media", "comment_media", "save_media", "unsave_media",
        "get_likers", "get_comment_replies", "reply_to_comment", "edit_caption",
        # Friendships
        "follow_user", "block_user", "unblock_user", "mute_user", "unmute_user",
        "remove_follower", "get_pending_requests", "approve_request",
        "get_mutual_followers", "get_friendship_status",
        # Stories
        "get_stories", "get_story_viewers", "react_to_story",
        "create_highlight", "get_all_highlights",
        # DM
        "send_dm",
        # Account
        "get_my_account",
        # Feed
        "get_timeline", "get_saved_posts", "get_liked_posts",
        # Growth
        "get_non_followers", "get_fans", "unfollow_non_followers",
        "follow_hashtag_users",
        # Export (needs followers/following data)
        "export_followers_csv", "export_following_csv", "export_post_likers",
        # Notifications
        "get_notifications", "get_activity_counts",
        # Explore
        "explore_feed",
        # Location (needs session)
        "search_locations", "get_location_info", "get_nearby_locations",
        "get_followers", "get_following",
        # ── Phase 5: Auth & Session ───────────────────────
        "validate_session", "logout",
        # ── Phase 5: Insights ─────────────────────────────
        "get_account_insights", "get_media_insight",
        "get_business_info", "get_ads_accounts",
        # ── Phase 5: Audience ─────────────────────────────
        "find_lookalike_audience", "get_audience_overlap",
        "get_audience_insights",
        # ── Phase 5: A/B Testing ──────────────────────────
        "create_ab_test", "run_ab_test",
        "get_ab_results", "list_ab_tests",
        # ── Phase 5: Automation ───────────────────────────
        "auto_dm_new_followers", "auto_comment_hashtag",
        "auto_like_feed", "auto_like_hashtag",
        "auto_watch_stories", "get_action_log",
        # ── Phase 5: Scheduler ────────────────────────────
        "schedule_post", "schedule_story", "schedule_reel",
        "list_scheduled_jobs", "cancel_scheduled_job",
        # ── Phase 5: Monitor ──────────────────────────────
        "monitor_account", "unmonitor_account",
        "monitor_check_now", "get_monitor_events",
        "get_monitor_stats",
        # ── Phase 5: Bulk Download ────────────────────────
        "bulk_download_posts", "bulk_download_stories",
        "bulk_download_highlights", "bulk_download_everything",
        # ── Phase 5: Comment Manager ──────────────────────
        "manage_comments", "auto_reply_comments",
        "delete_spam_comments", "get_comment_sentiment",
    }

    def _get_filtered_tools(self) -> List[Dict]:
        """
        Dynamically filter tools based on login status.

        Anonymous mode: removes ~45 login-required tools → saves ~3000 tokens
        Login mode: returns all tools
        """
        from .providers.base import instaharvest_v2_TOOLS

        if self._is_logged_in:
            return instaharvest_v2_TOOLS  # All 110 tools

        # Anonymous — filter out login-required tools
        filtered = [
            tool for tool in instaharvest_v2_TOOLS
            if tool["name"] not in self._LOGIN_REQUIRED_TOOLS
        ]

        logger.info(
            f"Tool filter: {len(instaharvest_v2_TOOLS)} → {len(filtered)} "
            f"(anonymous mode, {len(instaharvest_v2_TOOLS) - len(filtered)} hidden)"
        )
        return filtered

    # ═══════════════════════════════════════════════════════════
    # HISTORY COMPRESSION — TOKEN OPTIMIZATION
    # ═══════════════════════════════════════════════════════════

    # Compress when history has more than this many non-system messages
    _COMPRESS_THRESHOLD = 10
    # Keep this many recent messages after compression
    _KEEP_RECENT = 4

    def _compress_history(self) -> None:
        """
        Compress conversation history to save tokens.

        When history exceeds _COMPRESS_THRESHOLD non-system messages,
        summarize old messages into a single compact entry.
        Keeps: system prompt + summary + last _KEEP_RECENT messages.

        This reduces token usage by ~40-60% on long conversations.
        """
        # Count non-system messages
        non_system = [m for m in self._history if m["role"] != "system"]
        if len(non_system) < self._COMPRESS_THRESHOLD:
            return

        # Split: system + old messages + recent messages
        system_msg = self._history[0]  # Always system
        # Find the split point: keep last _KEEP_RECENT messages
        keep_count = self._KEEP_RECENT
        old_messages = non_system[:-keep_count]
        recent_messages = non_system[-keep_count:]

        if not old_messages:
            return

        # Build summary of old messages
        summary = self._summarize_messages(old_messages)

        if summary:
            # Rebuild history: system + summary + recent
            self._history = [
                system_msg,
                {
                    "role": "user",
                    "content": (
                        f"[CONVERSATION SUMMARY — previous {len(old_messages)} messages]\n"
                        f"{summary}\n"
                        f"[END SUMMARY — continue from here]"
                    ),
                },
                {
                    "role": "assistant",
                    "content": "Understood. I have read the previous conversation summary. Let's continue.",
                },
            ] + recent_messages

            logger.info(
                f"History compressed: {len(non_system)} msgs → "
                f"summary + {len(recent_messages)} recent "
                f"(saved ~{len(old_messages) * 200} tokens)"
            )

    def _summarize_messages(self, messages: List[Dict]) -> str:
        """
        Use LLM to summarize a list of messages into a compact summary.
        Falls back to simple extraction if LLM call fails.
        """
        # Build a compact representation of messages for summarization
        lines = []
        for msg in messages:
            role = msg.get("role", "?")
            content = str(msg.get("content", ""))

            if role == "user":
                lines.append(f"User: {content[:200]}")
            elif role == "assistant":
                # Only include non-empty assistant messages
                if content.strip():
                    lines.append(f"AI: {content[:300]}")
            elif role == "tool":
                # Tool results — very compact
                name = msg.get("name", "tool")
                lines.append(f"[Tool {name}: {content[:150]}]")

        conversation_text = "\n".join(lines)

        # Try to use LLM for smart summarization
        try:
            summary_prompt = [
                {
                    "role": "user",
                    "content": (
                        "Summarize this conversation in 3-5 bullet points. "
                        "Focus on: what the user asked, what data was retrieved, "
                        "what actions were taken. Be very concise.\n\n"
                        f"{conversation_text[:2000]}"
                    ),
                }
            ]

            response = self._provider.generate(
                messages=summary_prompt,
                temperature=0.1,
            )

            if response.content and response.content.strip():
                return response.content.strip()

        except Exception as e:
            logger.warning(f"LLM summarization failed: {e}")

        # Fallback: simple extraction of user messages
        user_msgs = [
            msg.get("content", "")[:100]
            for msg in messages
            if msg.get("role") == "user" and msg.get("content", "").strip()
        ]
        if user_msgs:
            return "Previous requests:\n" + "\n".join(f"- {m}" for m in user_msgs)

        return ""

    # ═══════════════════════════════════════════════════════════
    # AGENT LOOP — THE CORE
    # ═══════════════════════════════════════════════════════════

    def _agent_loop(
        self,
        step_callback: Optional[Callable[[Dict[str, Any]], None]] = None,
    ) -> AgentResult:
        """
        Main agent loop: LLM → Tool → Result → LLM → ...

        Continues until LLM responds without tool calls or max_steps reached.
        """
        result = AgentResult()

        def _emit(event: Dict[str, Any]):
            """Safely emit a step event via the callback."""
            if step_callback:
                try:
                    step_callback(event)
                except Exception:
                    pass  # Never let callback errors break the loop

        for step in range(self._max_steps):
            result.steps = step + 1

            if self._verbose:
                safe_print(f"  {emoji('📍', '>')} Step {step + 1}...", end="")

            # Emit: thinking
            _emit({"type": "thinking", "step": step + 1, "message": f"Step {step + 1}: Analyzing..."})

            # Call LLM with filtered tools
            try:
                response = self._provider.generate(
                    messages=self._history,
                    tools=self._get_filtered_tools(),
                    temperature=0.1,
                )
            except Exception as e:
                result.error = f"Error communicating with AI: {e}"
                logger.error(result.error)
                _emit({"type": "error", "message": result.error})
                return result

            # No tool calls — check if text contains code that should be executed
            if not response.has_tool_calls:
                content = response.content or ""

                # AUTO-EXECUTE: If response has ```python code blocks,
                # extract and run them automatically (MALFORMED fallback scenario)
                extracted_code = self._extract_code_from_text(content)
                if extracted_code:
                    if self._verbose:
                        print(f" 🔧 auto-exec")

                    # Emit: code extracted
                    _emit({"type": "code", "content": extracted_code, "description": "Auto-extracted Python code"})

                    self._history.append({"role": "assistant", "content": content})

                    # Execute the extracted code
                    _emit({"type": "tool_call", "name": "run_instaharvest_v2_code", "arguments": {"description": "auto-extracted"}})
                    exec_result = self._handle_code_execution(
                        {"code": extracted_code, "description": "auto-extracted"},
                        step_callback=step_callback,
                    )

                    # Track results
                    result.code_executed = extracted_code
                    if isinstance(exec_result, ExecutionResult):
                        result.execution_result = exec_result

                    # Add tool-like result to history so LLM sees the output
                    result_str = str(exec_result)
                    if len(result_str) > 3000:
                        result_str = result_str[:3000] + "\n... (truncated)"

                    # Emit: tool result
                    _emit({"type": "tool_result", "name": "run_instaharvest_v2_code", "output": result_str[:2000],
                           "success": isinstance(exec_result, ExecutionResult) and exec_result.success})

                    self._history.append({
                        "role": "tool",
                        "name": "run_instaharvest_v2_code",
                        "content": result_str,
                    })

                    # Continue loop — LLM will see the code output
                    # and generate a proper final answer
                    continue

                # Normal final answer (no code blocks found)
                result.answer = content
                self._history.append({"role": "assistant", "content": content})

                if self._verbose:
                    print(" ✅")
                return result

            # If response has BOTH text content AND tool calls (Gemini behavior)
            # — save the text as potential answer
            if response.content and response.content.strip():
                result.answer = response.content

            if self._verbose:
                print(f" 🔧 {len(response.tool_calls)} tool")

            self._add_assistant_message(response)

            for tc in response.tool_calls:
                # Emit: tool call
                _emit({"type": "tool_call", "name": tc.name, "arguments": tc.arguments})

                # ── Tool usage analytics ──
                self._tool_stats[tc.name] = self._tool_stats.get(tc.name, 0) + 1

                tool_result = self._execute_tool(tc, step_callback=step_callback)

                # Add tool result to history
                self._add_tool_result(tc, tool_result)

                # Emit: tool result
                tool_output = str(tool_result)
                is_error = tool_output.startswith("Error") or tool_output.startswith("❌")
                _emit({"type": "tool_result", "name": tc.name, "output": tool_output[:2000],
                       "success": not is_error})

                # ── Dead loop detection ──
                if is_error:
                    error_key = f"{tc.name}:{tool_output[:100]}"
                    self._error_tracker[error_key] = self._error_tracker.get(error_key, 0) + 1

                    if self._error_tracker[error_key] >= self._max_identical_errors:
                        # Inject guidance to break the loop
                        self._history.append({
                            "role": "user",
                            "content": (
                                f"⚠️ SYSTEM: Tool '{tc.name}' failed {self._error_tracker[error_key]} times "
                                f"with the same error. Try a DIFFERENT approach or tool, "
                                f"or tell the user what went wrong and stop."
                            ),
                        })
                        logger.warning(
                            f"Dead loop detected: {tc.name} failed "
                            f"{self._error_tracker[error_key]}x — injecting guidance"
                        )
                        if self._verbose:
                            print(f"    ⚠️ Dead loop detected — switching strategy")
                        break  # Break this tool iteration, go back to LLM
                else:
                    # Success — clear error tracker for this tool
                    keys_to_clear = [k for k in self._error_tracker if k.startswith(f"{tc.name}:")]
                    for k in keys_to_clear:
                        del self._error_tracker[k]

                # Track code and files
                if tc.name == "run_instaharvest_v2_code":
                    result.code_executed = tc.arguments.get("code", "")
                    if isinstance(tool_result, ExecutionResult):
                        result.execution_result = tool_result
                elif tc.name in ("save_to_file", "create_chart", "write_file"):
                    filename = tc.arguments.get("filename", tc.arguments.get("path", ""))
                    if filename:
                        result.files_created.append(filename)
                        self._files_created.append(filename)
                elif tc.name == "download_media":
                    output_dir = tc.arguments.get("output_dir", "downloads")
                    result.files_created.append(output_dir)

        # Max steps reached — use last captured answer if available
        if not result.answer.strip():
            result.answer = "Warning: step limit reached. Result may be incomplete."
        return result

    # ═══════════════════════════════════════════════════════════
    # CODE EXTRACTION
    # ═══════════════════════════════════════════════════════════

    @staticmethod
    def _extract_code_from_text(text: str) -> str:
        """
        Extract Python code from markdown code blocks in LLM text.

        Looks for ```python ... ``` blocks and returns the code.
        Returns empty string if no code block found.
        """
        if not text or "```python" not in text:
            return ""

        import re
        # Find all ```python ... ``` blocks
        pattern = r'```python\s*\n(.*?)```'
        matches = re.findall(pattern, text, re.DOTALL)

        if not matches:
            return ""

        # Use the largest code block (most likely the main one)
        code = max(matches, key=len).strip()

        # Sanity check: code should be meaningful (not just comments)
        code_lines = [l for l in code.split('\n') if l.strip() and not l.strip().startswith('#')]
        if len(code_lines) < 1:
            return ""

        return code

    # ═══════════════════════════════════════════════════════════
    # TOOL EXECUTION
    # ═══════════════════════════════════════════════════════════

    def _execute_tool(
        self,
        tool_call: ToolCall,
        step_callback: Optional[Callable[[Dict[str, Any]], None]] = None,
    ) -> Any:
        """Execute a single tool call with permission check."""
        name = tool_call.name
        args = tool_call.arguments

        # Core tools (handled directly)
        if name == "run_instaharvest_v2_code":
            return self._handle_code_execution(args, step_callback=step_callback)
        elif name == "save_to_file":
            return self._handle_save_file(args)
        elif name == "ask_user":
            return self._handle_ask_user(args)

        # Specialized Instagram tools — direct API calls, no sandbox
        elif name == "get_profile":
            if self._verbose:
                print(f"    🔍 get_profile(@{args.get('username', '?')})")
            return TOOL_HANDLERS[name](args, ig=self._ig, cache=self._user_cache)

        elif name == "get_posts":
            if self._verbose:
                print(f"    📸 get_posts(@{args.get('username', '?')})")
            return TOOL_HANDLERS[name](args, ig=self._ig, cache=self._user_cache)

        elif name == "search_users":
            if self._verbose:
                print(f"    🔎 search_users('{args.get('query', '?')}')")
            return TOOL_HANDLERS[name](args, ig=self._ig)

        elif name == "get_user_info":
            if self._verbose:
                print(f"    👤 get_user_info(@{args.get('username', '?')})")
            return TOOL_HANDLERS[name](
                args, ig=self._ig,
                is_logged_in=self._is_logged_in,
                cache=self._user_cache,
            )
        # ─── Phase 2/3/4 Specialized Tools ─────────────────────────
        elif name in (
            # Phase 2: Social actions (login required)
            "follow_user", "get_followers", "get_following",
            "get_friendship_status", "like_media", "comment_media",
            "get_stories", "send_dm", "get_hashtag_info", "get_my_account",
            # Phase 3: Public Anonymous (no login)
            "get_user_id", "is_public", "exists",
            "get_feed", "get_all_posts", "get_reels",
            "get_comments", "get_highlights", "get_similar_accounts",
            "get_post_by_shortcode", "get_post_by_url", "get_media_urls",
            "get_hashtag_posts", "get_location_posts", "run_diagnostics",
            # Phase 4: Upload & Content Creation
            "upload_photo", "upload_video", "upload_reel",
            "upload_story_photo", "upload_story_video",
            "upload_carousel", "delete_media",
            # Phase 4: Advanced Media
            "get_likers", "save_media", "unsave_media",
            "get_comment_replies", "reply_to_comment", "edit_caption",
            # Phase 4: Advanced Friendships
            "block_user", "unblock_user", "mute_user", "unmute_user",
            "remove_follower", "get_pending_requests",
            "approve_request", "get_mutual_followers",
            # Phase 4: Stories Management
            "get_story_viewers", "react_to_story",
            "create_highlight", "get_all_highlights",
            # Phase 4: Growth
            "get_non_followers", "get_fans",
            "unfollow_non_followers", "follow_hashtag_users",
            # Phase 4: Export & Pipeline
            "export_followers_csv", "export_following_csv",
            "export_post_likers", "export_to_json",
            "save_to_sqlite", "save_to_jsonl",
            # Phase 4: Location
            "search_locations", "get_location_info", "get_nearby_locations",
            # Phase 4: Feed
            "get_timeline", "get_saved_posts", "get_liked_posts",
            # Phase 4: Users
            "get_full_profile", "parse_bio",
            # Phase 4: Hashtag Research
            "analyze_hashtag", "suggest_hashtags",
            # Phase 4: Notifications
            "get_notifications", "get_activity_counts",
            # Phase 4: Public Data Analytics
            "compare_profiles", "engagement_analysis", "build_report",
            # Phase 4: Advanced Search
            "search_hashtags", "search_places", "explore_feed",
        ):
            emoji_map = {
                "follow_user": "👥", "get_followers": "👥",
                "get_following": "👥", "get_friendship_status": "🤝",
                "like_media": "❤️", "comment_media": "💬",
                "get_stories": "📱", "send_dm": "✉️",
                "get_hashtag_info": "#️⃣", "get_my_account": "👤",
                "get_user_id": "🆔", "is_public": "🔓", "exists": "❓",
                "get_feed": "📰", "get_all_posts": "📸", "get_reels": "🎬",
                "get_comments": "💬", "get_highlights": "⭐", "get_similar_accounts": "👥",
                "get_post_by_shortcode": "🔗", "get_post_by_url": "🌐",
                "get_media_urls": "📎", "get_hashtag_posts": "#️⃣",
                "get_location_posts": "📍", "run_diagnostics": "🔬",
                "upload_photo": "📤", "upload_video": "📤", "upload_reel": "📤",
                "upload_story_photo": "📤", "upload_story_video": "📤",
                "upload_carousel": "📤", "delete_media": "🗑️",
                "get_likers": "❤️", "save_media": "🔖", "unsave_media": "🔖",
                "get_comment_replies": "💬", "reply_to_comment": "💬", "edit_caption": "✏️",
                "block_user": "🚫", "unblock_user": "✅", "mute_user": "🔇", "unmute_user": "🔊",
                "remove_follower": "👋", "get_pending_requests": "📩",
                "approve_request": "✅", "get_mutual_followers": "🤝",
                "get_story_viewers": "👁", "react_to_story": "😊",
                "create_highlight": "⭐", "get_all_highlights": "⭐",
                "get_non_followers": "👻", "get_fans": "🌟",
                "unfollow_non_followers": "👻", "follow_hashtag_users": "📈",
                "export_followers_csv": "💾", "export_following_csv": "💾",
                "export_post_likers": "💾", "export_to_json": "💾",
                "save_to_sqlite": "🗄️", "save_to_jsonl": "📄",
                "search_locations": "📍", "get_location_info": "📍", "get_nearby_locations": "📍",
                "get_timeline": "📰", "get_saved_posts": "🔖", "get_liked_posts": "❤️",
                "get_full_profile": "👤", "parse_bio": "📋",
                "analyze_hashtag": "#️⃣", "suggest_hashtags": "💡",
                "get_notifications": "🔔", "get_activity_counts": "📊",
                "compare_profiles": "📊", "engagement_analysis": "📈", "build_report": "📋",
                "search_hashtags": "#️⃣", "search_places": "📍", "explore_feed": "🔍",
            }
            if self._verbose:
                print(f"    {emoji_map.get(name, '🔧')} {name}({', '.join(f'{k}={v!r}' for k, v in list(args.items())[:2])})")
            return TOOL_HANDLERS[name](
                args, ig=self._ig, is_logged_in=self._is_logged_in,
            )

        elif name == "get_media_info":
            if self._verbose:
                print(f"    📄 get_media_info({args.get('media_id', '?')})")
            return TOOL_HANDLERS[name](args, ig=self._ig)

        # Extended tools (handled by tools.py)
        elif name in TOOL_HANDLERS:
            handler = TOOL_HANDLERS[name]

            # System tools emoji routing
            _sys_emoji = {
                "write_file": "📝", "append_to_file": "📝",
                "copy_file": "📋", "move_file": "📦",
                "delete_file": "🗑️", "file_exists": "❓",
                "get_file_info": "ℹ️",
                "get_working_directory": "📂", "create_directory": "📁",
                "list_directory": "📂", "find_files": "🔍",
                "save_session_data": "💾", "load_session_data": "💾",
                "list_sessions": "💾",
                "get_env_var": "🔧", "set_working_directory": "📂",
                "get_system_info": "🖥️",
            }

            if self._verbose:
                e = _sys_emoji.get(name, "🔧")
                print(f"    {e} {name}")

            # Some tools need the ig instance
            if name == "download_media":
                if not self._permissions.check(
                    "download.media",
                    f"Download: {args.get('url', '?')}"
                ):
                    return "❌ Permission denied by user"
                return handler(args, ig=self._ig)

            elif name == "http_request":
                if not self._permissions.check(
                    "http.request",
                    f"HTTP {args.get('method', 'GET')} {args.get('url', '?')}"
                ):
                    return "❌ Permission denied by user"
                return handler(args)

            else:
                # Universal dispatch — pass ig, login status, and cache
                # Handlers use **kw to accept only what they need
                return handler(
                    args,
                    ig=self._ig,
                    is_logged_in=self._is_logged_in,
                    cache=self._user_cache,
                )

        else:
            return f"Unknown tool: {name}"

    def _handle_code_execution(
        self,
        args: Dict,
        step_callback: Optional[Callable[[Dict[str, Any]], None]] = None,
    ) -> str:
        """Handle run_instaharvest_v2_code tool call."""
        code = args.get("code", "")
        description = args.get("description", "Code execution")

        if not code.strip():
            return "Error: empty code"

        # Permission check
        if not self._permissions.check_code_execution(code):
            return "Permission denied by user"

        if self._verbose:
            _e = emoji('\U0001f4bb', '>')
            safe_print(f"    {_e} Executing code: {description}")

        # Emit: code about to be executed
        if step_callback:
            try:
                step_callback({"type": "code", "content": code, "description": description})
            except Exception:
                pass

        # Execute in sandbox
        exec_result = self._executor.run(code)

        if exec_result.success:
            output = str(exec_result)
            logger.info(f"Code executed successfully ({exec_result.duration:.2f}s)")
            return output
        else:
            logger.warning(f"Code error: {exec_result.error}")
            return f"Error: {exec_result.error}"

    def _handle_save_file(self, args: Dict) -> str:
        """Handle save_to_file tool call."""
        filename = args.get("filename", "")
        content = args.get("content", "")

        if not filename:
            return "Error: filename not specified"

        # Permission check
        if not self._permissions.check(
            "export.save_file",
            f"Save to file: {filename} ({len(content)} chars)"
        ):
            return "Permission denied by user"

        try:
            # Security: only relative paths
            if os.path.isabs(filename) or ".." in filename:
                return "Error: only relative paths allowed (current directory only)"

            # Get full absolute path for user feedback
            full_path = os.path.abspath(filename)

            # Ensure parent directory exists
            parent_dir = os.path.dirname(full_path)
            if parent_dir and not os.path.exists(parent_dir):
                os.makedirs(parent_dir, exist_ok=True)

            ext = os.path.splitext(filename)[1].lower()

            # Excel format — use openpyxl
            if ext in (".xlsx", ".xls"):
                return self._save_as_excel(filename, content, full_path)

            # Text-based formats
            with open(filename, "w", encoding="utf-8") as f:
                f.write(content)

            if self._verbose:
                _e = emoji('\U0001f4e5', '>')
                safe_print(f"    {_e} Saved: {full_path}")

            return f"File saved successfully!\nPath: {full_path}\nSize: {len(content)} chars"
        except Exception as e:
            return f"Error: could not write to file: {e}"

    def _save_as_excel(self, filename: str, content: str, full_path: str) -> str:
        """Save content as Excel file using openpyxl."""
        try:
            import openpyxl
        except ImportError:
            return (
                "Error: openpyxl is not installed. "
                "Install it: pip install openpyxl"
            )

        try:
            import json

            # Try to parse content as JSON (list of dicts or dict)
            data = json.loads(content)

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Data"

            if isinstance(data, list) and data:
                if isinstance(data[0], dict):
                    # List of dicts — headers from keys
                    headers = list(data[0].keys())
                    ws.append(headers)
                    for row in data:
                        ws.append([row.get(h, "") for h in headers])
                else:
                    # List of values
                    for item in data:
                        ws.append([item] if not isinstance(item, list) else item)
            elif isinstance(data, dict):
                # Single dict — key/value columns
                ws.append(["Key", "Value"])
                for k, v in data.items():
                    ws.append([str(k), str(v)])

            wb.save(filename)

            if self._verbose:
                _e = emoji('\U0001f4e5', '>')
                safe_print(f"    {_e} Excel saved: {full_path}")

            return f"Excel file saved successfully!\nPath: {full_path}\nRows: {ws.max_row}"

        except json.JSONDecodeError:
            # Content is not JSON — save as plain text in cells
            wb = openpyxl.Workbook()
            ws = wb.active
            for line in content.split("\n"):
                ws.append([line])
            wb.save(filename)
            return f"Excel file saved (text mode)!\nPath: {full_path}\nRows: {ws.max_row}"
        except Exception as e:
            return f"Error saving Excel: {e}"

    def _handle_ask_user(self, args: Dict) -> str:
        """Handle ask_user tool call."""
        question = args.get("question", "")
        try:
            answer = input(f"\nAgent asks: {question}\nYour answer: ").strip()
            return answer or "(no response)"
        except (KeyboardInterrupt, EOFError):
            return "(user did not respond)"

    # ═══════════════════════════════════════════════════════════
    # MESSAGE HISTORY MANAGEMENT
    # ═══════════════════════════════════════════════════════════

    def _add_assistant_message(self, response: ProviderResponse) -> None:
        """Add assistant message with tool calls to history."""
        provider_name = self._provider.__class__.__name__

        if "OpenAI" in provider_name or "Compatible" in provider_name:
            from .providers.openai_provider import OpenAIProvider
            msg = OpenAIProvider.format_assistant_with_tools(
                response.content, response.tool_calls
            )
            self._history.append(msg)
        elif "Claude" in provider_name:
            # Claude tool results need tool_use_id
            msg = {"role": "assistant", "content": response.content or ""}
            self._history.append(msg)
        else:
            # Generic format for Gemini and others
            msg = {"role": "assistant", "content": response.content or ""}
            self._history.append(msg)

    def _add_tool_result(self, tool_call: ToolCall, result: Any) -> None:
        """Add tool result to history."""
        result_str = str(result)

        # Truncate very long results
        if len(result_str) > 3000:
            result_str = result_str[:3000] + "\n... (qisqartirildi)"

        provider_name = self._provider.__class__.__name__

        if "OpenAI" in provider_name or "Compatible" in provider_name:
            from .providers.openai_provider import OpenAIProvider
            self._history.append(
                OpenAIProvider.format_tool_result(tool_call.id, result_str)
            )
        elif "Claude" in provider_name:
            # Claude expects tool_result in user message
            self._history.append({
                "role": "tool",
                "tool_use_id": tool_call.id,
                "content": result_str,
            })
        else:
            # For Gemini: send as tool response
            self._history.append({
                "role": "tool",
                "name": tool_call.name,
                "content": result_str,
            })

    def _print_history(self) -> None:
        """Print conversation history."""
        for msg in self._history:
            if msg["role"] == "system":
                continue
            role = {"user": "👤", "assistant": "🤖", "tool": "🔧"}.get(msg["role"], "?")
            content = str(msg.get("content", ""))[:100]
            print(f"  {role} {content}")

    # ═══════════════════════════════════════════════════════════
    # UTILITIES
    # ═══════════════════════════════════════════════════════════

    @staticmethod
    def _resolve_api_key(provider: str) -> Optional[str]:
        """Try to resolve API key from environment."""
        return resolve_api_key(provider)

    def __repr__(self) -> str:
        return (
            f"<InstaAgent provider={self._provider.provider_name} "
            f"permission={self._permissions.level.value} "
            f"history={len(self._history)}>"
        )
